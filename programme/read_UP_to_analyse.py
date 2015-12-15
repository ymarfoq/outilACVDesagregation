def read_UP_to_analyse(filename, unit_converter, UP_meta_info, infrastructure_rescale):
	from openpyxl.reader.excel import load_workbook
	wb = load_workbook(filename)
	system_counter = 0
	systems = []
	
	while 1:
		system_counter += 1
		try:
			ws = wb.get_sheet_by_name('system ' + str(system_counter))
			system = {}
			header = 1
			for row in ws.rows:
				if header == 1:
					header = 0
				else:
					(UP, quantity, unit) = row
					UP = str(UP.value)
					quantity = float(quantity.value)
					unit = str(unit.value)
					quantity = quantity * unit_converter[unit][1]
					if UP_meta_info[UP]['Infrastructure'] == 'Yes':
						quantity = quantity * infrastructure_rescale
					system[UP] = quantity
			systems.append(system)
		except AttributeError:
			break
	
	return systems
 
def read_all_UP_to_analyse(filename, unit_converter, UP_meta_info, infrastructure_rescale):
	from openpyxl.reader.excel import load_workbook
	wb = load_workbook(filename)
	systems = []
	ws = wb.get_sheet_by_name('systems')
		
	for row in ws.rows:
		system = {}
		try:
			(UP, quantity, unit) = row
		except:
			print row
		UP = str(UP.value)
		quantity = float(quantity.value)
		unit = str(unit.value)
		quantity = quantity * unit_converter[unit][1]
		if UP_meta_info[UP]['Infrastructure'] == 'Yes':
			quantity = quantity * infrastructure_rescale
		system[UP] = quantity
		systems.append(system)
	return systems

def read_all_UP_to_analyse_csv(filename, unit_converter, UP_meta_info, infrastructure_rescale):
	import re
	f=open('all_systems_to_analyse.txt',"r")
	units=["m3","m3"]
	i=0
	systems = []
	while 1:
		try:
			row=f.readline()
			row=row[:-1]
			system = {}
			UP=row
			quantity=float(1)
			unit=str(units[i])
			i+=1
			quantity = quantity * unit_converter[unit][1]
			if UP_meta_info[UP]['Infrastructure'] == 'Yes':
				quantity = quantity * infrastructure_rescale
			system[UP] = quantity
			systems.append(system)
		except:
			print row
			break
			
	return systems
